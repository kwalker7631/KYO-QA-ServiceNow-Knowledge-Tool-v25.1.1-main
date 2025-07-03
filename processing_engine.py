# processing_engine.py
import shutil, time, json, openpyxl, re
from queue import Queue
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import *
from custom_exceptions import FileLockError
from data_harvesters import harvest_all_data
from file_utils import is_file_locked
from ocr_utils import extract_text_from_pdf, _is_ocr_needed

def clear_review_folder():
    if PDF_TXT_DIR.exists():
        for f in PDF_TXT_DIR.glob("*.txt"):
            try:
                f.unlink()
            except OSError as e:
                print(f"Error deleting review file {f}: {e}")

def get_cache_path(pdf_path):
    try:
        return CACHE_DIR / f"{pdf_path.stem}_{pdf_path.stat().st_size}.json"
    except FileNotFoundError:
        return CACHE_DIR / f"{pdf_path.stem}_unknown.json"

# --- UPDATED FUNCTION ---
def process_single_pdf(pdf_path, progress_queue, ignore_cache=False):
    # Ensure pdf_path is a Path object for consistency
    pdf_path = Path(pdf_path)
    filename = pdf_path.name
    cache_path = get_cache_path(pdf_path)

    # FIX: Announce which file is being processed for live feedback in the terminal
    progress_queue.put({"type": "log", "tag": "info", "msg": f"Processing: {filename}"})
    
    if not ignore_cache and cache_path.exists():
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            if "status" not in cached_data:
                raise KeyError
            
            progress_queue.put({"type": "log", "tag": "info", "msg": f"Loaded from cache: {filename}"})
            if cached_data.get("status") == "Needs Review":
                progress_queue.put({"type": "review_item", "data": cached_data.get("review_info")})
            progress_queue.put({"type": "file_complete", "status": cached_data.get("status")})
            if cached_data.get("ocr_used"):
                progress_queue.put({"type": "increment_counter", "counter": "ocr"})
            return cached_data
        except (json.JSONDecodeError, KeyError):
             progress_queue.put({"type": "log", "tag": "warning", "msg": f"Corrupt cache for {filename}. Reprocessing..."})

    progress_queue.put({"type": "status", "msg": filename, "led": "Queued"})
    
    # FIX: Pass the absolute string path to the OCR utility to prevent file open errors
    absolute_pdf_path = str(pdf_path.resolve())
    
    ocr_required = _is_ocr_needed(absolute_pdf_path)
    if ocr_required:
        progress_queue.put({"type": "status", "msg": filename, "led": "OCR"})
        progress_queue.put({"type": "increment_counter", "counter": "ocr"})
    
    extracted_text = extract_text_from_pdf(absolute_pdf_path)
    if not extracted_text.strip():
        result = {"filename": filename, "models": "Error: Text Extraction Failed", "author": "", "status": "Fail", "ocr_used": ocr_required, "review_info": None}
    else:
        progress_queue.put({"type": "status", "msg": filename, "led": "AI"})
        data = harvest_all_data(extracted_text, filename)
        if data["models"] == "Not Found":
            status = "Needs Review"
            review_txt_path = PDF_TXT_DIR / f"{pdf_path.stem}.txt"
            with open(review_txt_path, 'w', encoding='utf-8') as f:
                f.write(f"--- Filename: {filename} ---\n\n{extracted_text}")
            review_info = {"filename": filename, "reason": "No models", "txt_path": str(review_txt_path), "pdf_path": str(pdf_path)}
            progress_queue.put({"type": "review_item", "data": review_info})
        else:
            status = "Pass"
            review_info = None
        result = {"filename": filename, **data, "status": status, "ocr_used": ocr_required, "review_info": review_info}

    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(result, f)
    progress_queue.put({"type": "file_complete", "status": result["status"]})
    return result

def run_processing_job(job_info, progress_queue, cancel_event, pause_event):
    try:
        is_rerun = job_info.get("is_rerun", False)
        excel_path = Path(job_info["excel_path"])
        input_path = job_info["input_path"]
        progress_queue.put({"type": "log", "tag": "info", "msg": "Processing job started."})

        if is_rerun:
            clear_review_folder()
            cloned_path = excel_path
        else:
            ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            cloned_path = OUTPUT_DIR / f"cloned_{excel_path.stem}_{ts}{excel_path.suffix}"
            if is_file_locked(excel_path):
                raise FileLockError("Input Excel is locked.")
            shutil.copy(excel_path, cloned_path)
        
        files = [Path(f) for f in input_path] if isinstance(input_path, list) else list(Path(input_path).glob('*.pdf'))
        results = {}
        for i, path in enumerate(files):
            if cancel_event.is_set():
                break
            if pause_event and pause_event.is_set():
                progress_queue.put({"type": "status", "msg": "Paused", "led": "Paused"})
                while pause_event.is_set():
                    time.sleep(0.5)
            progress_queue.put({"type": "progress", "current": i + 1, "total": len(files)})
            res = process_single_pdf(path, progress_queue, ignore_cache=is_rerun)
            if res is None:
                res = process_single_pdf(path, progress_queue, ignore_cache=True)
            if res:
                results[res["filename"]] = res

        if cancel_event.is_set():
            progress_queue.put({"type": "finish", "status": "Cancelled"})
            return

        progress_queue.put({"type": "status", "msg": "Updating Excel...", "led": "Saving"})
        workbook = openpyxl.load_workbook(cloned_path)
        sheet = workbook.active
        headers = [c.value for c in sheet[1]]
        if STATUS_COLUMN_NAME not in headers:
            sheet.cell(row=1, column=len(headers) + 1).value = STATUS_COLUMN_NAME
            headers.append(STATUS_COLUMN_NAME)
        cols = {h: headers.index(h) + 1 for h in [DESCRIPTION_COLUMN_NAME, META_COLUMN_NAME, AUTHOR_COLUMN_NAME, STATUS_COLUMN_NAME]}
       
        for row in sheet.iter_rows(min_row=2):
            desc = str(row[cols[DESCRIPTION_COLUMN_NAME]-1].value)
            for filename, data in results.items():
                if Path(filename).stem in desc:
                    row[cols[META_COLUMN_NAME]-1].value = data["models"]
                    row[cols[AUTHOR_COLUMN_NAME]-1].value = data["author"]
                    row[cols[STATUS_COLUMN_NAME]-1].value = f"{data['status']}{' (OCR)' if data['ocr_used'] else ''}"
                    break
        
        progress_queue.put({"type": "status", "msg": "Applying formatting...", "led": "Saving"})
        fills = {
            "Pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "Fail": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "Needs Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "OCR": PatternFill(start_color="0A9BCD", end_color="0A9BCD", fill_type="solid")
        }
        
        for row in sheet.iter_rows(min_row=2):
            status_val = str(row[cols[STATUS_COLUMN_NAME]-1].value)
            fill_key = status_val.replace(" (OCR)", "").strip()
            fill = fills.get(fill_key)
            if fill:
                for cell in row:
                    cell.fill = fill
            if "(OCR)" in status_val:
                row[cols[STATUS_COLUMN_NAME]-1].fill = fills["OCR"]
        
        for i, col in enumerate(sheet.columns, 1):
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            sheet.column_dimensions[get_column_letter(i)].width = (max_len + 2) if max_len < 60 else 60

        workbook.save(cloned_path)
        progress_queue.put({"type": "result_path", "path": str(cloned_path)})
        progress_queue.put({"type": "finish", "status": "Complete"})

    except Exception as e:
        progress_queue.put({"type": "log", "tag": "error", "msg": f"Critical error: {e}"})
        progress_queue.put({"type": "finish", "status": f"Error: {e}"})